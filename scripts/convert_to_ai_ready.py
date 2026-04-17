from __future__ import annotations

import csv
import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = ROOT / "ai_ready"

SOURCES = [
    {
        "id": "install_list",
        "kind": "xlsx",
        "label": "엘리베이터TV 설치리스트(외부용)_260202",
        "path": ROOT / "엘리베이터TV 설치리스트(외부용)_260202.xlsx",
    },
    {
        "id": "proposal_sample",
        "kind": "xlsx",
        "label": "포커스미디어 제안서 샘플",
        "path": ROOT / "★포커스미디어 제안서 샘플★.xlsx",
    },
    {
        "id": "automation_request_ppt",
        "kind": "pptx",
        "label": "포커스미디어 오토마타 제작 의뢰",
        "path": ROOT / "tjp인공지능연구소_오토마타의뢰(0312).pptx",
    },
]

NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}


def is_non_empty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def normalize_text(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.replace("\r\n", "\n").replace("\r", "\n").strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def clean_header(value: Any, fallback: str) -> str:
    if not is_non_empty(value):
        return fallback
    text = str(value).replace("\n", " / ").strip()
    return text or fallback


def ensure_unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique: list[str] = []
    for header in headers:
        count = seen.get(header, 0)
        if count:
            unique.append(f"{header} ({count + 1})")
        else:
            unique.append(header)
        seen[header] = count + 1
    return unique


def detect_header_row(value_rows: list[list[Any]]) -> int | None:
    best_row = None
    best_score = (-1, -1)
    max_scan_rows = min(len(value_rows), 30)
    for row_idx in range(1, max_scan_rows + 1):
        values = value_rows[row_idx - 1]
        non_empty = [value for value in values if is_non_empty(value)]
        if len(non_empty) < 3:
            continue
        string_count = sum(isinstance(value, str) for value in non_empty)
        score = (len(non_empty), string_count)
        if score > best_score:
            best_row = row_idx
            best_score = score
    return best_row


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False))
            handle.write("\n")


def build_row_exports(
    value_rows: list[list[Any]], formula_rows: list[list[Any]]
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    min_col = None
    max_col = None
    non_empty_cells = 0

    for row_idx, (value_row, formula_row) in enumerate(zip(value_rows, formula_rows), start=1):
        cell_map: dict[str, Any] = {}
        formula_map: dict[str, str] = {}
        max_width = max(len(value_row), len(formula_row))
        for col_idx in range(1, max_width + 1):
            value = value_row[col_idx - 1] if col_idx - 1 < len(value_row) else None
            formula_value = formula_row[col_idx - 1] if col_idx - 1 < len(formula_row) else None
            has_formula = isinstance(formula_value, str) and formula_value.startswith("=")
            if not is_non_empty(value) and not has_formula:
                continue
            column_letter = get_column_letter(col_idx)
            cell_map[column_letter] = value
            if has_formula:
                formula_map[column_letter] = formula_value
            if min_col is None or col_idx < min_col:
                min_col = col_idx
            if max_col is None or col_idx > max_col:
                max_col = col_idx
            non_empty_cells += 1

        if cell_map:
            row_payload = {"row": row_idx, "cells": cell_map}
            if formula_map:
                row_payload["formulas"] = formula_map
            rows.append(row_payload)

    min_row = rows[0]["row"] if rows else None
    max_row = rows[-1]["row"] if rows else None

    summary = {
        "non_empty_row_count": len(rows),
        "non_empty_cell_count": non_empty_cells,
        "used_range": (
            f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            if rows and min_col and max_col and min_row and max_row
            else None
        ),
    }
    return rows, summary


def build_table_exports(
    value_rows: list[list[Any]], header_row: int | None
) -> tuple[list[dict[str, Any]], list[str], list[int]]:
    if header_row is None:
        return [], [], []

    header_columns: list[int] = []
    header_names: list[str] = []
    header_values = value_rows[header_row - 1]
    for col_idx in range(1, len(header_values) + 1):
        value = header_values[col_idx - 1]
        if not is_non_empty(value):
            continue
        header_columns.append(col_idx)
        header_names.append(clean_header(value, f"Column {col_idx}"))

    if len(header_columns) < 3:
        return [], [], []

    unique_headers = ensure_unique_headers(header_names)
    table_rows: list[dict[str, Any]] = []

    for row_idx in range(header_row + 1, len(value_rows) + 1):
        row_data = {"row_number": row_idx}
        row_has_value = False
        for col_idx, header in zip(header_columns, unique_headers):
            row_values = value_rows[row_idx - 1]
            value = row_values[col_idx - 1] if col_idx - 1 < len(row_values) else None
            row_data[header] = value
            if is_non_empty(value):
                row_has_value = True
        if row_has_value:
            table_rows.append(row_data)

    return table_rows, unique_headers, header_columns


def write_table_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    fieldnames = ["row_number", *headers]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def convert_xlsx(source: dict[str, Any]) -> dict[str, Any]:
    output_dir = OUTPUT_DIR / source["id"]
    output_dir.mkdir(parents=True, exist_ok=True)

    wb_values = load_workbook(source["path"], data_only=True, read_only=True)
    wb_formulas = load_workbook(source["path"], data_only=False, read_only=True)

    workbook_summary: dict[str, Any] = {
        "id": source["id"],
        "label": source["label"],
        "source_path": str(source["path"]),
        "kind": "xlsx",
        "sheets": [],
    }

    overview_lines = [
        f"# {source['label']}",
        "",
        f"- Source: `{source['path']}`",
        f"- Workbook type: `xlsx`",
        "",
    ]

    for index, sheet_name in enumerate(wb_values.sheetnames, start=1):
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]
        sheet_prefix = f"sheet_{index:02d}"

        value_rows = [
            [normalize_text(cell) for cell in row]
            for row in ws_values.iter_rows(values_only=True)
        ]
        formula_rows = [
            [cell for cell in row]
            for row in ws_formulas.iter_rows(values_only=True)
        ]

        row_exports, row_summary = build_row_exports(value_rows, formula_rows)
        header_row = detect_header_row(value_rows)
        table_rows, table_headers, table_columns = build_table_exports(value_rows, header_row)

        rows_jsonl = output_dir / f"{sheet_prefix}_rows.jsonl"
        write_jsonl(rows_jsonl, row_exports)

        sheet_summary = {
            "sheet_index": index,
            "sheet_name": sheet_name,
            "dimensions": {
                "max_row": len(value_rows),
                "max_column": max((len(row) for row in value_rows), default=0),
            },
            "row_export_file": rows_jsonl.name,
            "header_row": header_row,
            **row_summary,
        }

        if table_rows:
            csv_path = output_dir / f"{sheet_prefix}_table.csv"
            jsonl_path = output_dir / f"{sheet_prefix}_table.jsonl"
            write_table_csv(csv_path, table_headers, table_rows)
            write_jsonl(jsonl_path, table_rows)
            sheet_summary["table_export"] = {
                "columns": table_headers,
                "source_columns": [get_column_letter(col_idx) for col_idx in table_columns],
                "row_count": len(table_rows),
                "csv_file": csv_path.name,
                "jsonl_file": jsonl_path.name,
            }

        workbook_summary["sheets"].append(sheet_summary)

        overview_lines.extend(
            [
                f"## {index}. {sheet_name}",
                "",
                f"- Used range: `{row_summary['used_range']}`",
                f"- Non-empty rows: `{row_summary['non_empty_row_count']}`",
                f"- Non-empty cells: `{row_summary['non_empty_cell_count']}`",
                f"- Row export: `{rows_jsonl.name}`",
                f"- Header row guess: `{header_row}`",
            ]
        )
        if table_rows:
            overview_lines.extend(
                [
                    f"- Table rows: `{len(table_rows)}`",
                    f"- Table columns: `{', '.join(table_headers)}`",
                    "",
                ]
            )
        else:
            overview_lines.append("")

    write_json(output_dir / "workbook_summary.json", workbook_summary)
    (output_dir / "README.md").write_text("\n".join(overview_lines).strip() + "\n", encoding="utf-8")
    return workbook_summary


def extract_slide_texts(ppt_path: Path) -> list[dict[str, Any]]:
    slides: list[dict[str, Any]] = []
    with ZipFile(ppt_path) as archive:
        slide_names = sorted(
            name
            for name in archive.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        for index, slide_name in enumerate(slide_names, start=1):
            root = ET.fromstring(archive.read(slide_name))
            texts = []
            for node in root.findall(".//a:t", NS):
                text = (node.text or "").strip()
                if text:
                    texts.append(text)
            slides.append(
                {
                    "slide_number": index,
                    "source_xml": slide_name,
                    "text_items": texts,
                    "combined_text": " / ".join(texts),
                }
            )
    return slides


def convert_pptx(source: dict[str, Any]) -> dict[str, Any]:
    output_dir = OUTPUT_DIR / source["id"]
    output_dir.mkdir(parents=True, exist_ok=True)

    slides = extract_slide_texts(source["path"])
    summary = {
        "id": source["id"],
        "label": source["label"],
        "source_path": str(source["path"]),
        "kind": "pptx",
        "slide_count": len(slides),
        "slides_file": "slides.json",
    }

    write_json(output_dir / "slides.json", slides)

    md_lines = [
        f"# {source['label']}",
        "",
        f"- Source: `{source['path']}`",
        f"- Slide count: `{len(slides)}`",
        "",
    ]
    for slide in slides:
        md_lines.append(f"## Slide {slide['slide_number']}")
        md_lines.append("")
        if slide["text_items"]:
            for item in slide["text_items"]:
                md_lines.append(f"- {item}")
        else:
            md_lines.append("- (no extracted text)")
        md_lines.append("")

    (output_dir / "slides.md").write_text("\n".join(md_lines).strip() + "\n", encoding="utf-8")
    write_json(output_dir / "presentation_summary.json", summary)
    return summary


def build_root_readme(results: list[dict[str, Any]]) -> None:
    lines = [
        "# AI-Ready Exports",
        "",
        "원본 문서를 AI가 읽기 쉬운 텍스트 중심 구조로 변환한 결과물입니다.",
        "",
        "## Export formats",
        "",
        "- `*_rows.jsonl`: 시트의 비어 있지 않은 행을 Excel 열 문자 기준으로 보존한 행 단위 덤프",
        "- `*_table.csv`: 감지된 헤더 행을 기준으로 만든 구조화 테이블",
        "- `*_table.jsonl`: 같은 테이블의 JSONL 버전",
        "- `slides.md` / `slides.json`: PPT 슬라이드별 텍스트 추출 결과",
        "",
        "## Documents",
        "",
    ]

    for result in results:
        lines.append(f"- `{result['id']}`: `{result['label']}`")
    lines.append("")

    (OUTPUT_DIR / "README.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    results: list[dict[str, Any]] = []

    for source in SOURCES:
        if source["kind"] == "xlsx":
            results.append(convert_xlsx(source))
        elif source["kind"] == "pptx":
            results.append(convert_pptx(source))
        else:
            raise ValueError(f"Unsupported source kind: {source['kind']}")

    build_root_readme(results)
    write_json(OUTPUT_DIR / "manifest.json", results)
    print(f"Exported {len(results)} source documents to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
