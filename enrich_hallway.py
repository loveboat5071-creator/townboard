import openpyxl
import requests
import time
import re
import os

# 공공데이터포털 설정 (사용자 확인 키)
SERVICE_KEY = '500c1df5c162639d9a9dd87f3b8bc4f5c81d57d7331eadb7c8e38bcf77f05215'

# 서비스 엔드포인트
LIST_URL = 'https://apis.data.go.kr/1613000/AptListService3/getSigunguAptList3'
INFO_URL = 'https://apis.data.go.kr/1613000/AptBasisInfoServiceV4/getAphusBassInfoV4'

# 시군구 코드 매핑 (주요 지역 우선)
SIGUNGU_MAP = {
    ('서울', '마포'): '11440', ('서울', '서초'): '11650', ('서울', '강남'): '11680',
    ('서울', '송파'): '11710', ('서울', '용산'): '11170', ('서울', '영등포'): '11560',
    ('서울', '관악'): '11620', ('서울', '강동'): '11740', ('서울', '강서'): '11500',
    ('서울', '구로'): '11530', ('서울', '금천'): '11545', ('서울', '동작'): '11590',
    ('서울', '양천'): '11470', ('서울', '광진'): '11215', ('서울', '동대문'): '11230',
    ('서울', '성동'): '11200', ('서울', '성북'): '11290', ('서울', '은평'): '11380',
    ('서울', '서대문'): '11410', ('서울', '노원'): '11350', ('서울', '도봉'): '11320',
    ('서울', '강북'): '11300', ('서울', '중랑'): '11260', ('서울', '중'): '11140', ('서울', '종로'): '11110',
    ('경기', '과천'): '41290', ('경기', '성남'): '41130', ('경기', '용인'): '41460',
    ('경기', '수원'): '41110', ('경기', '안양'): '41170', ('경기', '부천'): '41190',
    ('경기', '광명'): '41210', ('경기', '평택'): '41220', ('경기', '이천'): '41500',
    ('경기', '안산'): '41270', ('경기', '고양'): '41280', ('경기', '구리'): '41310',
    ('경기', '남양주'): '41360', ('경기', '의정부'): '41150', ('경기', '하남'): '41450',
    ('인천', '연수'): '28185', ('인천', '남동'): '28200', ('인천', '서'): '28260',
    ('대구', '수성'): '27260', ('대구', '중'): '27110', ('대구', '달서'): '27290',
    ('부산', '해운대'): '26350', ('부산', '수영'): '26500', ('부산', '남'): '26290'
}

apt_cache = {}

def get_apt_list_in_district(sigungu_code):
    if sigungu_code in apt_cache:
        return apt_cache[sigungu_code]
    
    try:
        params = {'serviceKey': SERVICE_KEY, 'sigunguCode': sigungu_code, 'numOfRows': 2000, 'pageNo': 1}
        res = requests.get(LIST_URL, params=params, timeout=15)
        # XML 파싱 (단지명: 코드)
        items = re.findall(r'<item>.*?<kaptCode>(.*?)</kaptCode>.*?<kaptName>(.*?)</kaptName>.*?</item>', res.text, re.DOTALL)
        mapping = {name.strip(): code for code, name in items}
        apt_cache[sigungu_code] = mapping
        return mapping
    except Exception as e:
        print(f"\n[오류] 목록 조회 실패 ({sigungu_code}): {e}")
        return {}

def get_hallway_type(kapt_code):
    try:
        params = {'serviceKey': SERVICE_KEY, 'kaptCode': kapt_code}
        res = requests.get(INFO_URL, params=params, timeout=10)
        match = re.search(r'<kaptRtTypeNm>([^<]+)</kaptRtTypeNm>', res.text)
        return match.group(1) if match else "정보없음"
    except:
        return "오류"

def main():
    input_file = 'input.xlsx'
    output_file = 'ᄐ아운보드_가동리스트_복도유형_완성본.xlsx'
    
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # 컬럼 위치 자동 감지
    header = [str(cell.value or "").strip() for cell in ws[6]]
    try:
        city_col = header.index("지역1") + 1
        dist_col = header.index("지역2") + 1
        name_col = header.index("단지명") + 1
    except ValueError:
        # 단지명 대신 아파트명일 수 있음
        name_col = next((i+1 for i, v in enumerate(header) if "아파트" in v or "단지" in v), 4)
        city_col, dist_col = 6, 7

    target_col = ws.max_column + 1
    ws.cell(row=6, column=target_col).value = "복도유형"
    
    total = ws.max_row - 6
    print(f"🚀 총 {total}개 단지 분석 시작 (통합 지능형 엔진)...")

    for i in range(7, ws.max_row + 1):
        city = str(ws.cell(row=i, column=city_col).value or "").strip()
        dist = str(ws.cell(row=i, column=dist_col).value or "").strip()
        name = str(ws.cell(row=i, column=name_col).value or "").strip()
        
        if not name: continue
        
        # 1. 시군구 코드 찾기
        sigungu_code = SIGUNGU_MAP.get((city, dist))
        if not sigungu_code:
            # 구 이름만으로 재검색 (서울 강남 -> 강남)
            for (c, d), code in SIGUNGU_MAP.items():
                if d == dist:
                    sigungu_code = code
                    break
        
        print(f"[{i-6}/{total}] {city} {dist} {name}...", end=" ", flush=True)
        
        if sigungu_code:
            dist_list = get_apt_list_in_district(sigungu_code)
            
            # 2. 이름 매칭 (부분 일치)
            matched_code = None
            # 정확히 일치 우선
            if name in dist_list:
                matched_code = dist_list[name]
            else:
                # 포함 관계 조사
                for kname, kcode in dist_list.items():
                    if name in kname or kname in name:
                        matched_code = kcode
                        break
            
            if matched_code:
                h_type = get_hallway_type(matched_code)
                ws.cell(row=i, column=target_col).value = h_type
                print(f"성공! ({h_type})")
            else:
                ws.cell(row=i, column=target_col).value = "미발견"
                print("미발견")
        else:
            ws.cell(row=i, column=target_col).value = "지역코드없음"
            print("지원정지구역")
            
        if i % 100 == 0:
            wb.save(output_file)
        
        time.sleep(0.05) # 대량 조회 최적화

    wb.save(output_file)
    print(f"\n✨ 모든 분석 완료! 완성된 파일을 확인해 주세요: {output_file}")

if __name__ == "__main__":
    main()
