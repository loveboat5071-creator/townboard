import openpyxl

def inspect_excel(filename):
    wb = openpyxl.load_workbook(filename, read_only=True)
    ws = wb.active
    print(f"Total Rows: {ws.max_row}")
    
    for i in range(1, 21):
        row_values = [cell.value for cell in ws[i]]
        # 텍스트가 있는 첫 번째 행을 헤더 후보로 간주
        if any(v is not None for v in row_values):
            print(f"Row {i}: {row_values}")

inspect_excel('input.xlsx')
